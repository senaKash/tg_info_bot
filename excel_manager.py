import openpyxl
import random
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from typing import List

class ExcelManager:
    def __init__(self, file_name):
        self.file_name = file_name
        self.list_names = ["ASP_Net_Core_data", "Lekcii_URLS", "Materiali_URLS"]

    def create_or_update_workbook(self, id, name):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            if self.list_names[0] not in workbook.sheetnames:
                self.create_template(workbook)
            self.add_data(workbook, id, name)
            workbook.save(self.file_name)
            print("Workbook updated successfully.")
        except FileNotFoundError:
            print("File not found. Creating a new workbook.")
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # Удаляем автоматически созданный пустой лист
            self.create_template(workbook)
            self.add_data(workbook, id, name)
            workbook.save(self.file_name)
            print(f"New workbook '{self.file_name}' created with template.")

    def create_template(self, workbook):
        for sheet_name in self.list_names:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
            if sheet_name == self.list_names[0]:  # Добавляем шапку только в первый лист
                sheet["A1"] = "id"
                sheet["B1"] = "имя"
                sheet["C1"] = "посещаемость"



    def add_data(self, workbook, id, name): #посещаемость, менять в падлу
        sheet = workbook[self.list_names[0]]

        # Проверяем, существует ли уже такой ID в столбце A
        id_column = sheet['A']
        id_exists = any(cell.value == id for cell in id_column)

        if not id_exists:
            max_row = sheet.max_row
            new_id = max_row if max_row > 1 else 1  # Generate unique id
            sheet.append([id, name, 0])
        else:
            print(f"ID {id} already exists in the workbook.")

    def check_and_update_attendance(self, id):
        now = datetime.now() + timedelta(hours=3)
        print(now)
        if now.weekday() == 6 and 17 <= now.hour < 20:
            try:
                workbook = openpyxl.load_workbook(self.file_name)
                sheet = workbook.active

                # Формируем название нового столбца
                column_name = now.strftime("%d.%m")

                column_exists = any(cell.value == column_name for cell in sheet[1])

                # Проверяем, существует ли столбец с заданным значением
                column_exists = False
                for cell in sheet[1]:
                    if cell.value == column_name:
                        column_index = cell.column  # Получаем индекс столбца
                        column_exists = True
                        break

                # Если столбец не существует, создаем новый
                if not column_exists:
                    column_letter = get_column_letter(sheet.max_column + 1)
                    sheet[column_letter + '1'] = column_name
                    workbook.save(self.file_name)
                    print(f"New column '{column_name}' added to the workbook.")
                    column_index = sheet.max_column  # Получаем индекс нового столбца

                id_exists = False
                for cell in sheet['A']:
                    if cell.value == id:
                        id_exists = True
                        break
                if id_exists:
                    #print(cell.row, column_index)
                    attendance_cell = sheet.cell(row=cell.row, column=column_index)
                    fill = attendance_cell.fill
                    if fill.fill_type == "solid":
                        print(f"Cell {attendance_cell.coordinate} is already green.")
                        return 2  # Возвращаем 2, чтобы показать, что ячейка уже была закрашена
                    else:
                        # Закрашиваем ячейку зеленым цветом
                        attendance_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        workbook.save(self.file_name)
                        print(f"Cell {attendance_cell.coordinate} is now green.")
                        return 1  # Возвращаем 1, чтобы показать, что ячейка была успешно закрашена
                else:
                    print(f"ID {id} not found in the workbook.")
                    return 0  # Возвращаем 0, чтобы показать, что ячейка не была найдена
            except FileNotFoundError:
                print("File not found.")
        else:
            print("Current time is not within the specified interval (Sunday 15:00 - 20:00).")
            return 3  # Возвращаем 0, чтобы показать, что не удалось проверить и обновить посещаемос

    def get_all_ids(self) -> List[int]:
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            if self.list_names[0] in workbook.sheetnames:
                sheet = workbook[self.list_names[0]]
                ids = []
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
                    if row[0] is not None:
                        ids.append(row[0])
                return ids
            else:
                print(f"List named '{self.list_names[0]}' not found in workbook.")
                return []
        except FileNotFoundError:
            print("Workbook not found.")
            return []


    def lektsionnie_url(self, url: str = None):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            if self.list_names[1] not in workbook.sheetnames:
                workbook.create_sheet(self.list_names[1])
            sheet = workbook[self.list_names[1]]

            if url is not None:
                # Добавление URL в первый свободный столбец A
                max_row = sheet.max_row + 1
                sheet[f"A{max_row}"] = url
                workbook.save(self.file_name)
                return None  # Возвращаем None, потому что задача была в добавлении URL
            else:
                # Чтение всех URL из столбца A и их возврат
                urls = []
                for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                    if row[0] is not None:  # Проверяем, что строка не пустая
                        urls.append(row[0])
                return urls
        except FileNotFoundError:
            print("Workbook not found.")
            return None